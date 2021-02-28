from tkinter import *
from tkinter import filedialog as fd
from PIL import ImageTk, Image
import openpyxl

def txtfile_opener():
    global txt
    txt = fd.askopenfilename()
def xlsfile_opener():
    global file
    file = fd.askopenfilename()
def retrieve_input():
    global inputValue
    inputValue=textBox.get("1.0","end-1c")
def popupmsg():
    popup = Tk()
    popup.geometry('450x300')
    popup.wm_title("Successful")
    label = Label(popup, text="Successfully Marked Attendence",fg='red')
    label.grid(row=1,column=1,pady=20,padx=30)
    label.config(font=('Times', 20,'bold'))
    B1 = Button(popup, text="Okay",padx=10,pady=10,command = popup.destroy)
    B1.grid(row=3,column=1,pady=20,padx=30)
    popup.mainloop()
def attendence():
    f = open(txt, "r")
    book = openpyxl.load_workbook(file)
    sheet = book.active
    plus=0
    imd=0
    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=1000, max_col=1):
        for cell in row:
            if cell.value != None:
                plus+=1
            else:
                imd=1
        if imd == 1:
            break
    c = 0
    for row in sheet.iter_cols(min_row=1, min_col=2, max_row=1, max_col=32):
        for cell in row:
            if cell.value == None:
                d = cell
                c = 1
        if c == 1:
            break
    d.value = inputValue
    i = 2
    while i <= plus:
        z = 0
        for row in sheet.iter_rows(min_row=i, min_col=1, max_row=i, max_col=32):
            for cell in row:
                if z == 0:
                    s = cell.value
                    z += 1
                else:
                    if cell.value == None:
                        x = cell
                        c = 1
                        break
            if c == 1:
                break
        count = 0
        for line in f:
            if (line.find(s) != -1):
                count += 1
        if (count > 0):
            x.value = 'P'
        else:
            x.value = 'A'
        f.seek(0, 0)
        i += 1
    book.save(file)
    popupmsg()
base = Tk()
base.geometry('1500x900')
base.title("EasymeeT")
txt=""
xls=""
inputValue=""
img = ImageTk.PhotoImage(Image.open("easymeet.png"))
base.iconphoto(False, img)
panel = Label(base, image = img)
panel.grid(row=0,column =3, padx = 200, pady = 10)
l1 = Label(base, text="Attendence marker using chat transcript",pady=20,padx=10,fg='black')
l1.config(font=("Times", 20))
l1.grid(row=1,column =3, padx = 10, pady = 10)
l4 = Label(base, text="Enter Date",padx=10,fg='black')
l4.config(font=("Times", 20))
l4.grid(row=2,column=3, padx = 10)
textBox=Text(base, height=2, width=40)
textBox.grid(row=3,column=3, padx = 10, pady = 20)
buttonCommit=Button(base,height=1, width=30,text="Confirm Date",bg="black",fg="red",
                    command=lambda: retrieve_input())
buttonCommit.grid(row=4,column=3, padx = 10, pady = 10)
button1 = Button(base,width=30, text ='Select a .txt/.sbv file',bg="black",fg="red", command = lambda:txtfile_opener())
button1.grid(row=5,column=3, padx = 10, pady = 30)
button2 = Button(base,width=30, text ='Select a .xls/.xlsx file',bg="black",fg="red", command = lambda:xlsfile_opener())
button2.grid(row=6,column=3, padx = 10, pady = 20)
button3 = Button(base,text='Submit',width=30,bg="black",fg="red",command=attendence)
button3.grid(row = 7, column = 3, padx = 10, pady = 20)
l2 = Label(base, text="Developed by Harshit Shukla",pady=20,fg='black')
l2.config(font=("Times", 20))
l2.grid(row=8,column=2, padx = 10, pady = 10)
mainloop()