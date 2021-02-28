import openpyxl
f = open("demo.sbv", "r")
book = openpyxl.load_workbook('sheet.xlsx')
sheet = book.active
c=0
for row in sheet.iter_cols(min_row=1, min_col=2, max_row=1, max_col=32):
        for cell in row:
            if cell.value == None:
                d=cell
                c=1
        if c==1:
            break
d.value=input('Enter the date')
i=2
while i<9:
    z=0
    for row in sheet.iter_rows(min_row=i, min_col=1, max_row=i, max_col=32):
        for cell in row:
            if z==0 :
                s = cell.value
                z+=1
            else:
                if cell.value == None:
                    x = cell
                    c = 1
                    break
        if c == 1:
             break
    count=0
    for line in f:
        if(line.find(s)!= -1):
            count+=1
    if(count>0):
        x.value ='P'
    else :
        x.value ='A'
    f.seek(0, 0)
    i+=1
book.save('sheet.xlsx')